"""
Excel output.

One sheet per page of the source document, each holding that page's table with its
hierarchy intact - indentation for nesting, bold for subtotals. Pages are kept apart
rather than merged, because a page can carry a different table from the one before it
and stitching them together produces something nobody can use.

Then two sheets that make the numbers defensible:

  Exceptions  only the checks that failed - the review worklist
  Audit       every model inference and every correction applied
  Provenance  every figure, the page and row it was printed on, and what proves it

The page sheets are the deliverable and are left exactly as the document presented them.
Everything added to make the numbers defensible goes on its own sheet, appended after
them, so opening the workbook looks the same as it always did.

The parser's raw response is not dumped here. It is retained in blob storage, because
for a rent roll it runs to thousands of single-cell rows that nobody reads.
"""

from datetime import date
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

from . import __version__, provenance
from .config import settings
from .models import (
    ROLE_BLOCK_HEADER, ROLE_SECTION, ROLE_SUBTOTAL, Document, Table, normalize_label,
)
from .reconcile import Check, summarize

NAVY = "1F3864"
HDR = PatternFill("solid", start_color=NAVY, end_color=NAVY)
SUB = PatternFill("solid", start_color="D9E1F2", end_color="D9E1F2")
OK_FILL = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
BAD_FILL = PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4")
INFO_FILL = PatternFill("solid", start_color="EDF3FA", end_color="EDF3FA")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
MUTED = Font(size=9, italic=True, color="595959")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
NUM = '#,##0.00;[Red](#,##0.00)'

# Dates are displayed the way the source document writes them, so the workbook and the
# PDF read the same. Showing a US statement's 11/15/2017 back as 15/11/2017 is correct
# but reads as though something was changed.
DATE_DISPLAY = {"MDY": "mm/dd/yyyy", "DMY": "dd/mm/yyyy", "YMD": "yyyy-mm-dd"}
DEFAULT_DATEFMT = "dd/mm/yyyy"


def _widths(ws, widths: Dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _header_row(ws, row: int, labels: List[str], fill=HDR, font=HDR_FONT) -> None:
    for j, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=j, value=label)
        cell.fill, cell.font, cell.border = fill, font, THIN
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _write(cell, value: Any, datefmt: str = DEFAULT_DATEFMT) -> None:
    """Write a value with the format its type deserves."""
    if value is None:
        return
    cell.value = value
    if isinstance(value, bool):
        return
    if isinstance(value, date):
        cell.number_format = datefmt
    elif isinstance(value, (int, float)):
        cell.number_format = NUM
    else:
        cell.alignment = Alignment(horizontal="left")


def _layout_groups(document: Document) -> Dict[str, List[Table]]:
    from .rules import fingerprint

    groups: Dict[str, List[Table]] = {}
    for table in document.tables:
        groups.setdefault(fingerprint(table), []).append(table)
    return groups


def _is_repeated_header(row, table: Table) -> bool:
    """A page header repeated on every page of a multi-page report."""
    key = normalize_label(row.label)
    return bool(key) and not row.has_figures and key == normalize_label(table.label_header)


def _signature(table: Table) -> tuple:
    """
    What makes two pages 'the same table'.

    Compares heading TEXT, not column count: two unrelated tables can be the same width,
    and joining those would silently file one table's figures under another's headings.
    Normalised so a difference of case or spacing does not split a table.
    """
    return (normalize_label(table.label_header),) + tuple(
        normalize_label(c) for c in table.column_names
    )


def _runs(tables: List[Table]) -> List[List[Table]]:
    """
    Group pages into runs of CONSECUTIVE pages sharing a signature.

    Consecutive matters. If pages 1-2 hold one table, page 3 another, and page 4 returns
    to the first shape, page 4 starts a third sheet rather than being appended to pages
    1-2 - a table that resumes after a different one is a separate table, and merging
    them would put rows in an order the document never had.
    """
    ordered = sorted(tables, key=lambda t: t.page)
    groups: List[List[Table]] = []
    for table in ordered:
        if not table.rows:
            continue  # an empty page earns no sheet
        if groups and _signature(groups[-1][-1]) == _signature(table):
            groups[-1].append(table)
        else:
            groups.append([table])
    return groups


def _make_filterable(ws, headings: List[str], last_row: int, last_col: int,
                     table_id: int) -> None:
    """
    Give the header row filter dropdowns and make the range sortable.

    An autofilter is applied in every case, because it tolerates the duplicate headings
    the parser sometimes produces ("Unit Type Resident Name" twice, before that fix is
    approved). Where the headings are already unique the range is also registered as a
    real Excel table, which adds banded rows and lets formulas reference columns by name.
    Excel rejects a table with duplicate headers, so that step is conditional rather than
    something that would corrupt the file.
    """
    if last_row < 2 or last_col < 1:
        return

    ref = f"A1:{get_column_letter(last_col)}{last_row}"
    cleaned = [str(h).strip() for h in headings]
    unique = (
        len({h.lower() for h in cleaned}) == len(cleaned)
        and all(cleaned)
    )

    # Exactly one of the two, never both. An Excel table carries its own filter, so
    # adding a worksheet autofilter over the same range defines the filter twice and
    # Excel opens the file in repair mode.
    if unique:
        try:
            # named per workbook rather than from a module-level counter, which two
            # concurrent requests could otherwise collide on
            table = XLTable(displayName=f"Extract{table_id}", ref=ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(table)
            return
        except Exception:
            pass  # fall through to a plain filter rather than lose the sheet

    ws.auto_filter.ref = ref


def sheet_map(document: Document) -> Dict[int, str]:
    """
    Which worksheet each page's figures land on.

    Built here rather than in the caller so the workbook and the provenance API cannot
    drift into naming the same sheet two different things - a citation that points at
    "Pages 2-3" when the workbook calls it "Page 2" is worse than no citation.
    """
    taken: set = set()
    pages: Dict[int, str] = {}
    for group in _runs(document.tables):
        name = _sheet_name(group, taken)
        for table in group:
            pages.setdefault(table.page, name)
    return pages


def _sheet_name(group: List[Table], taken: set) -> str:
    pages = [t.page for t in group]
    base = f"Page {pages[0]}" if len(pages) == 1 else f"Pages {pages[0]}-{pages[-1]}"
    name = base[:31]
    n = 2
    while name in taken:
        suffix = f" ({n})"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    taken.add(name)
    return name


# --------------------------------------------------------------------------
# One page, one sheet
# --------------------------------------------------------------------------

def _page_sheet(ws, document: Document, group: List[Table], table_id: int = 1) -> None:
    """
    Write one table, which may span several consecutive pages.

    Headings sit in row 1 so the sheet can be filtered or referenced directly, and the
    hierarchy is carried in the label column: nesting by indentation, subtotals in bold.
    When the table spans pages a Page column is added, so every row still traces back to
    where it was printed.
    """
    first = group[0]
    multi = len(group) > 1
    datefmt = DATE_DISPLAY.get(
        ((document.meta or {}).get("locale") or {}).get("date_order"), DEFAULT_DATEFMT
    )
    offset = 1 if multi else 0

    headings = (["Page"] if multi else []) + \
               [first.label_header or "Line item"] + list(first.column_names)
    _header_row(ws, 1, headings)
    ws.row_dimensions[1].height = 30

    r = 2
    for table in group:
        for row in table.rows:
            if not row.label and not row.has_figures:
                continue
            if _is_repeated_header(row, table):
                continue  # written once, in row 1

            if multi:
                page_cell = ws.cell(row=r, column=1, value=table.page)
                page_cell.border = THIN
                page_cell.alignment = Alignment(horizontal="center")

            label_cell = ws.cell(row=r, column=1 + offset,
                                 value=("    " * row.depth) + row.label)
            label_cell.border = THIN
            if row.role == ROLE_SUBTOTAL:
                label_cell.font = Font(bold=True)
                label_cell.fill = SUB
            elif row.role in (ROLE_SECTION, ROLE_BLOCK_HEADER):
                label_cell.font = Font(bold=True, color=NAVY)

            for j, value in enumerate(row.display(first.ncols)):
                c = ws.cell(row=r, column=2 + offset + j)
                c.border = THIN
                _write(c, value, datefmt)
                if row.role == ROLE_SUBTOTAL:
                    c.font = Font(bold=True)
                    c.fill = SUB
            r += 1

    last_row = r - 1
    last_col = len(headings)
    _make_filterable(ws, headings, last_row, last_col, table_id)

    ws.freeze_panes = ws.cell(row=2, column=2 + offset)
    widths = {1 + offset: 46, **{2 + offset + j: 16 for j in range(first.ncols)}}
    if multi:
        widths[1] = 7
    _widths(ws, widths)


# --------------------------------------------------------------------------
# Exceptions - the worklist
# --------------------------------------------------------------------------

def _exceptions_sheet(ws, checks: List[Check], state: Dict[str, Any]) -> None:
    summary = summarize(checks)
    failures = [c for c in checks if not c.passed]

    ws["A1"] = "Exceptions to review"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"{summary['passed']} of {summary['total']} checks tie. "
        + ("Nothing needs manual verification."
           if not failures else f"{len(failures)} item(s) below need a human.")
    )
    ws["A2"].font = Font(size=11, bold=True,
                         color="375623" if not failures else "9C0006")
    ws["A3"] = (
        "Only failures are listed. Every other figure has been verified against the "
        "document's own arithmetic, so review is limited to this list rather than every "
        "number in the report."
    )
    ws["A3"].font = MUTED

    _header_row(ws, 5, ["Check", "What was compared", "Document states",
                        "Figures add to", "Difference", "Where to look"])
    r = 6
    if not failures:
        ws.cell(row=r, column=1, value="none")
        ws.cell(row=r, column=2, value="every check tied")
        for j in range(1, 7):
            ws.cell(row=r, column=j).fill = OK_FILL
            ws.cell(row=r, column=j).border = THIN
        r += 1
    else:
        for check in failures:
            ws.cell(row=r, column=1, value=check.kind)
            ws.cell(row=r, column=2, value=check.description)
            _write(ws.cell(row=r, column=3), check.printed)
            _write(ws.cell(row=r, column=4), check.computed)
            _write(ws.cell(row=r, column=5), check.delta)
            ws.cell(row=r, column=6, value=check.location)
            for j in range(1, 7):
                ws.cell(row=r, column=j).fill = BAD_FILL
                ws.cell(row=r, column=j).border = THIN
            r += 1

    findings = state.get("findings") or []
    if findings:
        r += 1
        _header_row(ws, r, ["Structural finding", "Impact", "", "", "", ""])
        r += 1
        for finding in findings:
            ws.cell(row=r, column=1, value=finding.get("description"))
            ws.cell(row=r, column=2, value=finding.get("impact"))
            r += 1

    proposals = state.get("proposals") or []
    if proposals:
        r += 1
        _header_row(ws, r, ["Fix awaiting approval", "Rows", "Why", "", "", ""])
        r += 1
        for proposal in proposals:
            ws.cell(row=r, column=1, value=proposal.get("description"))
            ws.cell(row=r, column=2, value=proposal.get("affected_rows"))
            ws.cell(row=r, column=3, value=proposal.get("evidence"))
            r += 1

    # the exception list is a worklist, so it filters and sorts like one
    if failures:
        ws.auto_filter.ref = f"A5:F{5 + len(failures)}"
    ws.freeze_panes = "A6"
    _widths(ws, {1: 16, 2: 62, 3: 17, 4: 17, 5: 14, 6: 26})


# --------------------------------------------------------------------------
# Audit
# --------------------------------------------------------------------------

def _audit_sheet(ws, document: Document, state: Dict[str, Any]) -> None:
    ws["A1"] = "Audit trail"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        "Everything below is either a model inference or a change made to the parser's "
        "output. Figures were never read or altered by a model."
    )
    ws["A2"].font = MUTED

    meta = document.meta or {}
    _header_row(ws, 4, ["Model inference", "Value"])
    r = 5
    for key in ("company", "report_title", "statement_type", "period_label",
                "period_start", "period_end", "currency", "units", "basis",
                "segmentation", "notes", "_inferred_by"):
        if key not in meta or meta.get(key) in (None, ""):
            continue
        ws.cell(row=r, column=1, value=key.lstrip("_").replace("_", " "))
        ws.cell(row=r, column=2, value=str(meta[key]))
        r += 1

    for split in meta.get("column_headings_split") or []:
        ws.cell(row=r, column=1, value="column heading split")
        ws.cell(row=r, column=2,
                value=f"{split.get('merged')} -> {split.get('left')} | {split.get('right')}")
        r += 1

    # How figures and dates were read. Settled internally, recorded here so the decision
    # can be checked without anyone having been asked a question at upload time.
    locale_info = meta.get("locale") or {}
    if locale_info:
        r += 1
        _header_row(ws, r, ["How this document was read", "Basis"])
        r += 1
        for label, key in (("Figures", "numbers"), ("Dates", "dates")):
            if locale_info.get(key):
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=str(locale_info[key]))
                r += 1
        counts = (f"{locale_info.get('figures_read', 0)} figure(s) and "
                  f"{locale_info.get('dates_read', 0)} date(s) read")
        ws.cell(row=r, column=1, value="Read")
        ws.cell(row=r, column=2, value=counts)
        r += 1
        for conflict in locale_info.get("conflicts") or []:
            ws.cell(row=r, column=1, value="disagreed with the default")
            ws.cell(row=r, column=2, value=conflict)
            r += 1

    r = _change_log(ws, r + 1, document, state)
    _widths(ws, {1: 30, 2: 74, 3: 13, 4: 30, 5: 30})


def _change_log(ws, r: int, document: Document, state: Dict[str, Any]) -> int:
    """
    Every change made to the parser's output, and what permitted it.

    The question this answers is the one asked months later, by a client or by internal
    audit: did software alter this file, and on whose authority? Recording the change
    alone does not answer it. Recording the authority and the proof alongside does, and
    it is the difference between an agent that can be deployed on client work and one
    that cannot.

    Authority is one of three, and the distinction is the whole point:
      AUTO     the document's own arithmetic proves it; no permission needed
      PROPOSE  unprovable, so it was applied only because a human approved it once
      REPORT   described, never applied
    """
    _header_row(ws, r, ["Change", "What changed", "Authority", "Proof", "Approved"])
    r += 1

    rows: List[tuple] = []
    for x in state.get("repairs") or []:
        rows.append((
            "arithmetic repair",
            x.get("description"),
            str(x.get("authority") or "auto").upper(),
            x.get("proof") or x.get("evidence") or "",
            "",                      # proven, so nobody had to allow it
            x.get("at") or "",
        ))
    for x in state.get("applied_rules") or []:
        approver = x.get("approved_by") or ""
        approved_at = (x.get("approved_at") or "")[:10]
        rows.append((
            "approved correction",
            f"{x.get('description')} ({x.get('rows_changed', 0)} rows)",
            str(x.get("authority") or "propose").upper(),
            x.get("proof") or "",
            " ".join(p for p in (approver, approved_at) if p) or "approved previously",
            x.get("at") or "",
        ))

    if not rows:
        rows.append(("none", "the parser's output needed no correction", "", "", "", ""))

    for change, what, authority, proof, approved, at in sorted(rows, key=lambda t: t[5]):
        ws.cell(row=r, column=1, value=change)
        ws.cell(row=r, column=2, value=what)
        cell = ws.cell(row=r, column=3, value=authority)
        cell.font = MUTED if authority != "PROPOSE" else Font(bold=True)
        ws.cell(row=r, column=4, value=proof)
        ws.cell(row=r, column=5, value=approved)
        r += 1

    # What produced the file, so the output can be tied back to the build that made it.
    r += 1
    _header_row(ws, r, ["Produced by", "Value"])
    r += 1
    for label, value in (
        ("agent version", __version__),
        ("model", f"{settings.openai_deployment} - labels only, never figures"),
        ("parser", getattr(document, "parsing_engine", None) or "Evalueserve IDP"),
        ("run at", state.get("finished_at") or ""),
    ):
        if not value:
            continue
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=str(value))
        r += 1
    return r


# --------------------------------------------------------------------------
# Provenance - one row per figure, and what proves it
# --------------------------------------------------------------------------

PROVENANCE_HEADINGS = [
    "Page", "Sheet", "Row label", "Column", "As printed", "Value",
    "Status", "Out by", "Verified by", "Repaired",
]

_STATUS_FILL = {
    provenance.TIES: OK_FILL,
    provenance.EXCEPTION: BAD_FILL,
    provenance.UNCHECKED: INFO_FILL,
}

_STATUS_LABEL = {
    provenance.TIES: "ties",
    provenance.EXCEPTION: "exception",
    provenance.UNCHECKED: "not covered",
}


def _provenance_sheet(ws, document: Document, checks: List[Check],
                      sheet_of_page: Dict[int, str]) -> None:
    """
    Every figure, with the page it was printed on and the arithmetic that proves it.

    This sheet is the answer to "how do you know?" asked one figure at a time. It exists
    because the alternative - trusting the extraction because it looks right - is what
    an analyst is being paid to avoid, and because a figure that cannot name its source
    is indistinguishable from one a model invented.

    'As printed' is deliberately the raw string rather than the parsed number. A reviewer
    sent to find -1234.56 on a page that reads (1,234.56) will conclude the extraction is
    wrong; sent to find (1,234.56), they find it immediately.
    """
    citations = provenance.build(document, checks, sheet_of_page=sheet_of_page)
    summary = provenance.summarize(citations)

    ws["A1"] = "Provenance - every figure, and where it came from"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"{summary['cited']} of {summary['figures']} figures trace back to a page, a row "
        f"and the text the document printed. {summary['ties']} verified by arithmetic, "
        f"{summary['exceptions']} exception(s), {summary['unchecked']} not covered by a check."
    )
    ws["A2"].font = Font(size=11, bold=True,
                         color="9C0006" if summary["exceptions"] else "375623")
    ws["A3"] = (
        "No figure on this sheet was produced by a model. Each was moved by "
        "deterministic code from the parser's output, so it can still be traced to the "
        "characters the source document used. 'Not covered' means no check applies to "
        "that figure - it is not a pass."
    )
    ws["A3"].font = MUTED

    _header_row(ws, 5, PROVENANCE_HEADINGS)

    r = 6
    for c in citations:
        ws.cell(row=r, column=1, value=c.page)
        ws.cell(row=r, column=2, value=c.sheet or "")
        ws.cell(row=r, column=3, value=c.row_label)
        ws.cell(row=r, column=4, value=c.column)
        # forced to text: Excel would otherwise read "(1,234.56)" as a formula-ish
        # negative and re-render it, which is precisely the string being preserved
        printed = ws.cell(row=r, column=5, value=c.printed)
        printed.number_format = "@"
        printed.alignment = Alignment(horizontal="left")
        _write(ws.cell(row=r, column=6), c.value)
        ws.cell(row=r, column=7, value=_STATUS_LABEL.get(c.status, c.status))
        _write(ws.cell(row=r, column=8), c.out_by)
        ws.cell(row=r, column=9, value="; ".join(c.checks) if c.checks else "")
        ws.cell(row=r, column=10, value="; ".join(c.repairs) if c.repairs else "")

        fill = _STATUS_FILL.get(c.status, INFO_FILL)
        for j in range(1, len(PROVENANCE_HEADINGS) + 1):
            cell = ws.cell(row=r, column=j)
            cell.border = THIN
            if j == 7:
                cell.fill = fill
        r += 1

    # A plain autofilter, never an Excel table: the same range carrying both is what
    # made Excel open an earlier build in repair mode.
    if citations:
        ws.auto_filter.ref = f"A5:{get_column_letter(len(PROVENANCE_HEADINGS))}{r - 1}"
    ws.freeze_panes = "A6"
    _widths(ws, {1: 7, 2: 14, 3: 34, 4: 16, 5: 15, 6: 15,
                 7: 13, 8: 12, 9: 52, 10: 24})


# --------------------------------------------------------------------------

def build(document: Document, checks: List[Check], state: Dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    # consecutive pages sharing a heading signature form one table, one sheet
    sheet_of_page = sheet_map(document)
    for n, group in enumerate(_runs(document.tables), start=1):
        _page_sheet(wb.create_sheet(sheet_of_page[group[0].page]), document, group,
                    table_id=n)

    _exceptions_sheet(wb.create_sheet("Exceptions"), checks, state)
    _audit_sheet(wb.create_sheet("Audit"), document, state)
    _provenance_sheet(wb.create_sheet("Provenance"), document, checks, sheet_of_page)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
